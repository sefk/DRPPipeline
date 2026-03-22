"""Tests for sheet Claimed-column tally (XLSX path)."""

import io
import unittest

from openpyxl import Workbook

from utils.sheet_claimed_tally import (
    ClaimedTallyReport,
    find_claimed_columns_with_header_skips,
    find_named_column_index_and_skip,
    header_cell_matches_claimed,
    tally_claimed_from_xlsx_bytes,
)


class TestSheetClaimedTally(unittest.TestCase):
    def _assert_url_tally_skipped(self, result: ClaimedTallyReport) -> None:
        self.assertEqual(result.unclaimed_url_rows_by_sheet, ())
        self.assertEqual(result.sheets_without_url_column, ())
        self.assertIsNone(result.url_column_name)

    def test_header_cell_matches_claimed_word(self) -> None:
        self.assertTrue(header_cell_matches_claimed("Claimed"))
        self.assertTrue(header_cell_matches_claimed("Person who claimed"))
        self.assertTrue(header_cell_matches_claimed("Status (claimed)"))
        self.assertTrue(header_cell_matches_claimed("CLAIMED BY"))
        self.assertFalse(header_cell_matches_claimed(""))
        self.assertFalse(header_cell_matches_claimed("unclaimed"))
        self.assertFalse(header_cell_matches_claimed("disclaimed"))
        self.assertFalse(header_cell_matches_claimed("Foo"))

    def test_find_claimed_columns_with_header_skips(self) -> None:
        row1 = ["URL", "Claimed (add your name)", "Other", "x"]
        row2 = ["", "Person who claimed", "", "CLAIMED BY"]
        cols, skips = find_claimed_columns_with_header_skips(row1, row2)
        self.assertEqual(cols, [1, 3])
        self.assertEqual(skips[1], 2)
        self.assertEqual(skips[3], 2)

    def test_find_claimed_row1_only(self) -> None:
        row1 = ["URL", "Claimed", "Other"]
        cols, skips = find_claimed_columns_with_header_skips(row1, None)
        self.assertEqual(cols, [1])
        self.assertEqual(skips[1], 1)

    def test_find_claimed_row2_only(self) -> None:
        row1 = ["Topic", "URL"]
        row2 = ["x", "Who claimed"]
        cols, skips = find_claimed_columns_with_header_skips(row1, row2)
        self.assertEqual(cols, [1])
        self.assertEqual(skips[1], 2)

    def test_find_named_column_index_and_skip(self) -> None:
        row1 = ["Topic", "URL"]
        row2 = ["x", "Who claimed"]
        self.assertEqual(find_named_column_index_and_skip(row1, row2, "URL"), (1, 1))
        self.assertEqual(find_named_column_index_and_skip(row1, row2, "url"), (1, 1))
        self.assertIsNone(find_named_column_index_and_skip(row1, row2, "Dataset URL"))

    def test_find_named_column_row2_only(self) -> None:
        row1 = ["A", "B"]
        row2 = ["x", "URL"]
        self.assertEqual(find_named_column_index_and_skip(row1, row2, "URL"), (1, 2))

    def test_tally_claimed_from_xlsx_bytes_multi_sheet(self) -> None:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "TabA"
        ws1.append(["URL", "Claimed (add your name)", "Other"])
        ws1.append(["u1", "Alice", ""])
        ws1.append(["u2", "Bob", ""])
        ws2 = wb.create_sheet("TabB")
        ws2.append(["CLAIMED", "Note"])
        ws2.append(["Alice", "dup across tab"])
        ws2.append(["", "skip empty"])
        wb.create_sheet("NoClaimCol").append(["URL", "Other"])
        buf = io.BytesIO()
        wb.save(buf)
        result = tally_claimed_from_xlsx_bytes(buf.getvalue(), url_column_name="URL")
        self.assertEqual(result.tally["Alice"], 2)
        self.assertEqual(result.tally["Bob"], 1)
        self.assertEqual(result.total_claimed_entries, 3)
        self.assertEqual(result.unique_claimant_count, 2)
        self.assertIn("NoClaimCol", result.sheets_without_claimed_column)
        self.assertIn("TabB", result.sheets_without_url_column)
        self.assertEqual(dict(result.unclaimed_url_rows_by_sheet), {"TabA": 0})
        self.assertEqual(result.url_column_name, "URL")

    def test_tally_counts_multiple_claimed_columns_same_row(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["Claimed A", "Claimed B"])
        ws.append(["Same", "Same"])
        buf = io.BytesIO()
        wb.save(buf)
        result = tally_claimed_from_xlsx_bytes(buf.getvalue())
        self.assertEqual(result.tally["Same"], 2)
        self.assertEqual(result.total_claimed_entries, 2)
        self.assertEqual(result.unique_claimant_count, 1)
        self.assertEqual(result.sheets_without_claimed_column, ())
        self._assert_url_tally_skipped(result)

    def test_claimed_label_on_row2_data_starts_row3(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["Topic", "URL"])
        ws.append(["x", "Person who claimed"])
        ws.append(["a", "Zoe"])
        buf = io.BytesIO()
        wb.save(buf)
        result = tally_claimed_from_xlsx_bytes(buf.getvalue())
        self.assertEqual(result.tally.get("Zoe", 0), 1)
        self.assertNotIn("Person who claimed", result.tally)
        self._assert_url_tally_skipped(result)

    def test_empty_sheet_has_no_claimed_column(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "EmptyTab"
        buf = io.BytesIO()
        wb.save(buf)
        result = tally_claimed_from_xlsx_bytes(buf.getvalue())
        self.assertEqual(result.tally, {})
        self.assertEqual(result.total_claimed_entries, 0)
        self.assertEqual(result.unique_claimant_count, 0)
        self.assertEqual(result.sheets_without_claimed_column, ("EmptyTab",))
        self._assert_url_tally_skipped(result)

    def test_sheet_with_claimed_column_but_only_empty_cells(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "HasCol"
        ws.append(["Claimed"])
        ws.append([""])
        ws.append([None])
        buf = io.BytesIO()
        wb.save(buf)
        result = tally_claimed_from_xlsx_bytes(buf.getvalue())
        self.assertEqual(dict(result.tally), {})
        self.assertEqual(result.total_claimed_entries, 0)
        self.assertEqual(result.unique_claimant_count, 0)
        self.assertEqual(result.sheets_without_claimed_column, ())
        self._assert_url_tally_skipped(result)

    def test_unclaimed_url_rows_count_by_sheet(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inv"
        ws.append(["URL", "Claimed"])
        ws.append(["http://open.com", ""])
        ws.append(["http://taken.com", "Pat"])
        ws.append(["", ""])
        buf = io.BytesIO()
        wb.save(buf)
        result = tally_claimed_from_xlsx_bytes(buf.getvalue(), url_column_name="URL")
        self.assertEqual(dict(result.unclaimed_url_rows_by_sheet), {"Inv": 1})
        self.assertEqual(result.sheets_without_url_column, ())

    def test_unclaimed_url_multi_sheet(self) -> None:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "One"
        ws1.append(["URL", "Claimed"])
        ws1.append(["http://a", ""])
        ws1.append(["http://b", ""])
        ws2 = wb.create_sheet("Two")
        ws2.append(["URL", "Who claimed"])
        ws2.append(["http://c", ""])
        buf = io.BytesIO()
        wb.save(buf)
        result = tally_claimed_from_xlsx_bytes(buf.getvalue(), url_column_name="URL")
        self.assertEqual(dict(result.unclaimed_url_rows_by_sheet), {"One": 2, "Two": 1})


if __name__ == "__main__":
    unittest.main()
